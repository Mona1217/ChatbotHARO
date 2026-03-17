# app.py
from collections import deque
from datetime import datetime, timedelta
import hashlib
import hmac
from io import BytesIO
import json
import os
import re
import logging
from queue import Queue
import sqlite3
from dataclasses import dataclass, field
from threading import Condition, Lock, Thread
from typing import Dict, List, Optional, Tuple
from urllib.parse import urlparse

import requests
from dotenv import load_dotenv
from flask import Flask, Response, jsonify, redirect, render_template, request, send_file, session, url_for

load_dotenv()
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, "templates"),
    static_folder=os.path.join(BASE_DIR, "static"),
    static_url_path="/static",
)

# ========= Logging =========
logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO").upper(),
    format="%(asctime)s %(levelname)s %(name)s - %(message)s",
)
logger = app.logger  # usa el logger de Flask


def normalize_origin_value(origin: str) -> str:
    cleaned = (origin or "").strip()
    if cleaned == "*":
        return cleaned
    return cleaned.rstrip("/")


def resolve_app_path(path_value: str, fallback_name: str) -> str:
    cleaned = (path_value or "").strip()
    if not cleaned:
        cleaned = fallback_name
    if os.path.isabs(cleaned):
        return cleaned
    return os.path.join(BASE_DIR, cleaned)


# ========= Config =========
REQUIRED_ENV_VARS = (
    "WHATSAPP_VERIFY_TOKEN",
    "WHATSAPP_ACCESS_TOKEN",
    "WHATSAPP_PHONE_NUMBER_ID",
)

WHATSAPP_VERIFY_TOKEN = os.getenv("WHATSAPP_VERIFY_TOKEN", "")
WHATSAPP_ACCESS_TOKEN = os.getenv("WHATSAPP_ACCESS_TOKEN", "")
WHATSAPP_PHONE_NUMBER_ID = os.getenv("WHATSAPP_PHONE_NUMBER_ID", "")
WHATSAPP_APP_SECRET = os.getenv("WHATSAPP_APP_SECRET", "").strip()
WHATSAPP_API_VERSION = os.getenv("WHATSAPP_API_VERSION", "v24.0")
API_UNAVAILABLE_MESSAGE = os.getenv(
    "API_UNAVAILABLE_MESSAGE",
    "En este momento no estoy disponible, intenta mas tarde.",
)
BOT_PAUSED_MESSAGE = os.getenv(
    "BOT_PAUSED_MESSAGE",
    "En este momento no estoy disponible, intenta mas tarde.",
)
BOT_PAUSED = os.getenv("BOT_PAUSED", "false").strip().lower() in {"1", "true", "yes", "y"}
MONITOR_TOKEN = os.getenv("MONITOR_TOKEN", "").strip()
MONITOR_USERNAME = os.getenv("MONITOR_USERNAME", "").strip()
MONITOR_PASSWORD = os.getenv("MONITOR_PASSWORD", "").strip()
MONITOR_REFRESH_SECONDS = 5
MONITOR_DB_PATH = resolve_app_path(os.getenv("MONITOR_DB_PATH", ""), "monitor_events.db")
MONITOR_CORS_ORIGINS = {
    normalize_origin_value(origin)
    for origin in os.getenv(
        "MONITOR_CORS_ORIGINS",
        (
            "https://harorepositoty2-590358146556.europe-west1.run.app/,"
            "http://127.0.0.1:*,http://localhost:*,http://[::1]:*"
        ),
    ).split(",")
    if normalize_origin_value(origin)
}
MONITOR_CORS_PATHS = {
    "/monitor/events",
    "/api/monitor/events",
    "/monitor/pause",
    "/api/monitor/pause",
    "/monitor/stream",
    "/api/monitor/stream",
    "/monitor/export/contacts.xlsx",
    "/api/monitor/export/contacts.xlsx",
}
SESSION_COOKIE_SECURE = os.getenv("SESSION_COOKIE_SECURE", "").strip().lower() in {"1", "true", "yes", "y"}

try:
    # 0 o negativo => historial sin limite en memoria.
    MONITOR_LOG_MAX = int(os.getenv("MONITOR_LOG_MAX", "0"))
except ValueError:
    MONITOR_LOG_MAX = 0
try:
    SESSION_TIMEOUT_MINUTES = int(os.getenv("SESSION_TIMEOUT_MINUTES", "10"))
except ValueError:
    SESSION_TIMEOUT_MINUTES = 10
if SESSION_TIMEOUT_MINUTES <= 0:
    SESSION_TIMEOUT_MINUTES = 10
try:
    WEBHOOK_JOB_HISTORY_LIMIT = int(os.getenv("WEBHOOK_JOB_HISTORY_LIMIT", "1000"))
except ValueError:
    WEBHOOK_JOB_HISTORY_LIMIT = 1000
if WEBHOOK_JOB_HISTORY_LIMIT <= 0:
    WEBHOOK_JOB_HISTORY_LIMIT = 1000

USE_WELCOME_TEMPLATE = os.getenv("USE_WELCOME_TEMPLATE", "false").strip().lower() in {"1", "true", "yes", "y"}
WELCOME_TEMPLATE_NAME = os.getenv("WELCOME_TEMPLATE_NAME", "welcome_haro")
WELCOME_TEMPLATE_LANG = os.getenv("WELCOME_TEMPLATE_LANG", "es_CO")
CHATBOT_ENROLLMENT_WELCOME_IMAGE_URL = os.getenv("CHATBOT_ENROLLMENT_WELCOME_IMAGE_URL", "").strip()
CHATBOT_ENROLLMENT_WELCOME_IMAGE_CAPTION = os.getenv(
    "CHATBOT_ENROLLMENT_WELCOME_IMAGE_CAPTION",
    "",
).strip()

# ✅ Spring (tu backend real)
SPRING_BASE_URL = os.getenv(
    "SPRING_BASE_URL",
    "https://harorepositoty2-590358146556.europe-west1.run.app/",
).rstrip("/")
SPRING_AUTH_URL = os.getenv(
    "SPRING_AUTH_URL",
    f"{SPRING_BASE_URL}/api/auth/login",
).strip()
SPRING_AUTH_BODY_JSON = os.getenv("SPRING_AUTH_BODY_JSON", "").strip()
SPRING_AUTH_TOKEN_JSON_PATH = os.getenv("SPRING_AUTH_TOKEN_JSON_PATH", "token").strip() or "token"
SPRING_AUTH_HEADER_NAME = os.getenv("SPRING_AUTH_HEADER_NAME", "Authorization").strip() or "Authorization"
SPRING_AUTH_HEADER_PREFIX = os.getenv("SPRING_AUTH_HEADER_PREFIX", "Bearer ")

# ✅ Deduplicación para evitar reintentos duplicados
PROCESSED_MSG_IDS: set[str] = set()
MAX_PROCESSED_CACHE = 5000
SESSION_TIMEOUT = timedelta(minutes=SESSION_TIMEOUT_MINUTES)
SPRING_AUTH_TOKEN = ""
SPRING_AUTH_LOCK = Lock()
APP_SESSION_SECRET = (
    os.getenv("APP_SESSION_SECRET", "").strip()
    or os.getenv("SECRET_KEY", "").strip()
    or MONITOR_TOKEN
    or WHATSAPP_VERIFY_TOKEN
    or "chatbotharo-dev-secret"
)
app.secret_key = APP_SESSION_SECRET
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = SESSION_COOKIE_SECURE


# ========= Session State (solo si usas lógica local; puedes quitarlo si TODO vive en Spring) =========
@dataclass
class SessionState:
    step: str = "welcome"
    data: Dict[str, str] = field(default_factory=dict)
    last_activity: datetime = field(default_factory=datetime.now)


SESSIONS: Dict[str, SessionState] = {}
MESSAGE_EVENTS = deque(maxlen=MONITOR_LOG_MAX if MONITOR_LOG_MAX > 0 else None)
MESSAGE_EVENTS_LOCK = Lock()
MONITOR_DB_LOCK = Lock()
MONITOR_VERSION = 0
MONITOR_VERSION_COND = Condition()
WEBHOOK_QUEUE: Queue = Queue()
WEBHOOK_WORKER_LOCK = Lock()
WEBHOOK_WORKER_STARTED = False


# ========= Helpers =========
def validate_env() -> Tuple[bool, List[str]]:
    missing = [var for var in REQUIRED_ENV_VARS if not os.getenv(var)]
    return (len(missing) == 0, missing)


def normalize_text(text: str) -> str:
    return (text or "").strip().lower()


def clean_text_value(value: Optional[object]) -> str:
    if value is None:
        return ""
    return str(value).strip()


def optional_text_value(value: Optional[object]) -> Optional[str]:
    cleaned = clean_text_value(value)
    return cleaned or None


def is_restart_command(text: str) -> bool:
    return normalize_text(text) in {"menu", "inicio", "reiniciar", "start"}


def is_unrecognized_menu_warning(text: str) -> bool:
    normalized = normalize_text(text)
    return "no reconoc" in normalized and "opci" in normalized and "menu" in normalized


def is_valid_email(email: str) -> bool:
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", (email or "").strip()))


def parse_service_option(text: str) -> Optional[str]:
    value = normalize_text(text)
    mapping = {
        "1": "Ventas",
        "2": "Soporte tecnico",
        "3": "Agendar llamada",
        "ventas": "Ventas",
        "soporte": "Soporte tecnico",
        "soporte tecnico": "Soporte tecnico",
        "agendar": "Agendar llamada",
        "llamada": "Agendar llamada",
    }
    return mapping.get(value)


def normalize_event_text(value: Optional[str]) -> str:
    return clean_text_value(value)


def summarize_message_for_monitor(msg_type: Optional[str], text: Optional[str]) -> str:
    if text:
        return text
    return f"[{clean_text_value(msg_type) or 'unknown'}]"


def record_skipped_action(peer: str, action_type: str, reason: str) -> None:
    logger.warning("SKIPPED ACTION -> peer=%s type=%s reason=%s", peer, action_type, reason)
    add_monitor_event(
        direction="system",
        event_type="skipped_action",
        peer=peer,
        body=action_type or "unknown",
        status="skipped",
        detail=reason,
    )


def get_monitor_db_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(MONITOR_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_monitor_db() -> None:
    with MONITOR_DB_LOCK:
        with get_monitor_db_connection() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS monitor_events (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    version INTEGER NOT NULL UNIQUE,
                    ts TEXT NOT NULL,
                    direction TEXT NOT NULL,
                    event_type TEXT NOT NULL,
                    peer TEXT NOT NULL,
                    body TEXT NOT NULL,
                    status TEXT NOT NULL,
                    detail TEXT NOT NULL
                )
                """
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_monitor_events_version ON monitor_events(version)"
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_monitor_events_peer_version ON monitor_events(peer, version)"
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS processed_messages (
                    msg_id TEXT PRIMARY KEY,
                    created_at TEXT NOT NULL
                )
                """
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_processed_messages_created_at ON processed_messages(created_at)"
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS monitor_settings (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS webhook_jobs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source_label TEXT NOT NULL,
                    payload TEXT NOT NULL,
                    status TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    claimed_at TEXT NOT NULL DEFAULT '',
                    completed_at TEXT NOT NULL DEFAULT '',
                    attempts INTEGER NOT NULL DEFAULT 0,
                    last_error TEXT NOT NULL DEFAULT ''
                )
                """
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_webhook_jobs_status_id ON webhook_jobs(status, id)"
            )


def current_timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def get_last_monitor_ts_for_peer(peer: str) -> Optional[datetime]:
    try:
        with MONITOR_DB_LOCK:
            with get_monitor_db_connection() as conn:
                row = conn.execute(
                    """
                    SELECT ts
                    FROM monitor_events
                    WHERE peer = ?
                      AND direction = 'outbound'
                    ORDER BY version DESC
                    LIMIT 1
                    """,
                    (peer,),
                ).fetchone()
    except Exception as e:
        logger.exception("No se pudo leer monitor para peer=%s: %s", peer, e)
        return None

    if not row:
        return None

    raw_ts = row["ts"] if isinstance(row, sqlite3.Row) else row[0]
    if not raw_ts:
        return None

    try:
        return datetime.strptime(raw_ts, "%Y-%m-%d %H:%M:%S")
    except Exception:
        return None


def get_max_monitor_version_from_db() -> int:
    with MONITOR_DB_LOCK:
        with get_monitor_db_connection() as conn:
            row = conn.execute("SELECT COALESCE(MAX(version), 0) AS max_version FROM monitor_events").fetchone()
    return int(row["max_version"]) if row else 0


def insert_monitor_event_db(event: dict) -> None:
    with MONITOR_DB_LOCK:
        with get_monitor_db_connection() as conn:
            conn.execute(
                """
                INSERT OR REPLACE INTO monitor_events
                (version, ts, direction, event_type, peer, body, status, detail)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    int(event.get("version", 0)),
                    event.get("ts", ""),
                    event.get("direction", ""),
                    event.get("event_type", ""),
                    event.get("peer", ""),
                    event.get("body", ""),
                    event.get("status", ""),
                    event.get("detail", ""),
                ),
            )


def add_monitor_event(
    direction: str,
    event_type: str,
    peer: str = "",
    body: Optional[str] = None,
    status: str = "ok",
    detail: str = "",
) -> None:
    global MONITOR_VERSION

    with MONITOR_VERSION_COND:
        MONITOR_VERSION += 1
        version = MONITOR_VERSION

    normalized_body = normalize_event_text(body)
    if not normalized_body:
        normalized_body = f"[{clean_text_value(event_type) or 'event'}]"

    event = {
        "version": version,
        "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "direction": direction,
        "event_type": event_type,
        "peer": peer or "-",
        "body": normalized_body,
        "status": status,
        "detail": normalize_event_text(detail),
    }
    with MESSAGE_EVENTS_LOCK:
        MESSAGE_EVENTS.append(event)
    try:
        insert_monitor_event_db(event)
    except Exception as e:
        logger.exception("Error guardando evento en DB: %s", e)
    with MONITOR_VERSION_COND:
        MONITOR_VERSION_COND.notify_all()


def get_monitor_events(since: int = 0, peer: Optional[str] = None) -> List[dict]:
    where_clauses = []
    params: List[object] = []

    if since > 0:
        where_clauses.append("version > ?")
        params.append(since)

    if peer:
        where_clauses.append("peer = ?")
        params.append(peer)

    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""
    query = f"""
        SELECT version, ts, direction, event_type, peer, body, status, detail
        FROM monitor_events
        {where_sql}
        ORDER BY version ASC
    """

    try:
        with MONITOR_DB_LOCK:
            with get_monitor_db_connection() as conn:
                rows = conn.execute(query, params).fetchall()
        return [dict(row) for row in rows]
    except Exception as e:
        logger.exception("Error leyendo eventos de DB: %s", e)
        with MESSAGE_EVENTS_LOCK:
            events = list(MESSAGE_EVENTS)
        if peer:
            events = [ev for ev in events if ev.get("peer") == peer]
        if since > 0:
            events = [ev for ev in events if int(ev.get("version", 0)) > since]
        return events


def get_monitor_total_events(peer: Optional[str] = None) -> int:
    try:
        with MONITOR_DB_LOCK:
            with get_monitor_db_connection() as conn:
                if peer:
                    row = conn.execute(
                        "SELECT COUNT(*) AS total FROM monitor_events WHERE peer = ?",
                        (peer,),
                    ).fetchone()
                else:
                    row = conn.execute("SELECT COUNT(*) AS total FROM monitor_events").fetchone()
        return int(row["total"]) if row else 0
    except Exception as e:
        logger.exception("Error contando eventos de DB: %s", e)
        with MESSAGE_EVENTS_LOCK:
            events = list(MESSAGE_EVENTS)
        if peer:
            events = [ev for ev in events if ev.get("peer") == peer]
        return len(events)


def get_monitor_version() -> int:
    with MONITOR_VERSION_COND:
        return MONITOR_VERSION


def is_exportable_phone_peer(peer: str) -> bool:
    value = clean_text_value(peer)
    if not value or value in {"-", "Sistema"}:
        return False
    return bool(re.search(r"\d", value))


def normalize_phone_digits(peer: str) -> str:
    return re.sub(r"\D+", "", clean_text_value(peer))


def build_monitor_contact_rows_from_events(events: List[dict]) -> List[dict]:
    grouped: Dict[str, dict] = {}

    for event in events:
        peer = clean_text_value(event.get("peer"))
        if not is_exportable_phone_peer(peer):
            continue

        current = grouped.setdefault(
            peer,
            {
                "phone": peer,
                "phone_digits": normalize_phone_digits(peer),
                "first_ts": clean_text_value(event.get("ts")),
                "last_ts": clean_text_value(event.get("ts")),
                "total_events": 0,
                "inbound_events": 0,
                "outbound_events": 0,
                "system_events": 0,
                "last_version": int(event.get("version", 0) or 0),
            },
        )

        ts_value = clean_text_value(event.get("ts"))
        version_value = int(event.get("version", 0) or 0)
        direction_value = clean_text_value(event.get("direction"))

        if ts_value and (not current["first_ts"] or ts_value < current["first_ts"]):
            current["first_ts"] = ts_value
        if ts_value and (not current["last_ts"] or ts_value > current["last_ts"]):
            current["last_ts"] = ts_value

        current["total_events"] += 1
        current["last_version"] = max(current["last_version"], version_value)

        if direction_value == "inbound":
            current["inbound_events"] += 1
        elif direction_value == "outbound":
            current["outbound_events"] += 1
        else:
            current["system_events"] += 1

    return sorted(grouped.values(), key=lambda row: row["last_version"], reverse=True)


def get_monitor_contact_rows() -> List[dict]:
    query = """
        SELECT
            peer,
            MIN(ts) AS first_ts,
            MAX(ts) AS last_ts,
            COUNT(*) AS total_events,
            SUM(CASE WHEN direction = 'inbound' THEN 1 ELSE 0 END) AS inbound_events,
            SUM(CASE WHEN direction = 'outbound' THEN 1 ELSE 0 END) AS outbound_events,
            SUM(CASE WHEN direction = 'system' THEN 1 ELSE 0 END) AS system_events,
            MAX(version) AS last_version
        FROM monitor_events
        WHERE TRIM(peer) <> ''
          AND peer <> '-'
        GROUP BY peer
        ORDER BY last_version DESC
    """

    try:
        with MONITOR_DB_LOCK:
            with get_monitor_db_connection() as conn:
                rows = conn.execute(query).fetchall()
        raw_rows = [dict(row) for row in rows]
    except Exception as e:
        logger.exception("Error leyendo contactos de monitor desde DB: %s", e)
        with MESSAGE_EVENTS_LOCK:
            events = list(MESSAGE_EVENTS)
        return build_monitor_contact_rows_from_events(events)

    normalized_rows = []
    for row in raw_rows:
        peer = clean_text_value(row.get("peer"))
        if not is_exportable_phone_peer(peer):
            continue

        normalized_rows.append(
            {
                "phone": peer,
                "phone_digits": normalize_phone_digits(peer),
                "first_ts": clean_text_value(row.get("first_ts")),
                "last_ts": clean_text_value(row.get("last_ts")),
                "total_events": int(row.get("total_events", 0) or 0),
                "inbound_events": int(row.get("inbound_events", 0) or 0),
                "outbound_events": int(row.get("outbound_events", 0) or 0),
                "system_events": int(row.get("system_events", 0) or 0),
                "last_version": int(row.get("last_version", 0) or 0),
            }
        )

    return normalized_rows


def build_contacts_excel_file(rows: List[dict]) -> BytesIO:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("Falta openpyxl. Instala requirements.txt para exportar a Excel.") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Numeros bot"

    headers = [
        "Numero",
        "Numero normalizado",
        "Primer contacto",
        "Ultimo contacto",
        "Total eventos",
        "Mensajes entrantes",
        "Mensajes salientes",
        "Eventos de sistema",
    ]
    sheet.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row in rows:
        sheet.append(
            [
                row["phone"],
                row["phone_digits"],
                row["first_ts"],
                row["last_ts"],
                row["total_events"],
                row["inbound_events"],
                row["outbound_events"],
                row["system_events"],
            ]
        )

    for column_cells in sheet.columns:
        max_length = max(len(clean_text_value(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 14), 38)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def load_monitor_setting(key: str) -> Optional[str]:
    with MONITOR_DB_LOCK:
        with get_monitor_db_connection() as conn:
            row = conn.execute(
                "SELECT value FROM monitor_settings WHERE key = ? LIMIT 1",
                (key,),
            ).fetchone()
    if not row:
        return None
    return clean_text_value(row["value"] if isinstance(row, sqlite3.Row) else row[0])


def save_monitor_setting(key: str, value: str) -> None:
    with MONITOR_DB_LOCK:
        with get_monitor_db_connection() as conn:
            conn.execute(
                """
                INSERT INTO monitor_settings (key, value, updated_at)
                VALUES (?, ?, ?)
                ON CONFLICT(key) DO UPDATE SET
                    value = excluded.value,
                    updated_at = excluded.updated_at
                """,
                (key, clean_text_value(value), current_timestamp()),
            )


def load_runtime_settings() -> None:
    global BOT_PAUSED

    stored_bot_paused = load_monitor_setting("BOT_PAUSED")
    if stored_bot_paused is None:
        save_monitor_setting("BOT_PAUSED", "true" if BOT_PAUSED else "false")
        return

    BOT_PAUSED = stored_bot_paused.lower() in {"1", "true", "yes", "y"}


def create_webhook_job(source_label: str, payload: dict) -> int:
    payload_json = json.dumps(payload or {}, ensure_ascii=False)
    created_at = current_timestamp()
    with MONITOR_DB_LOCK:
        with get_monitor_db_connection() as conn:
            cursor = conn.execute(
                """
                INSERT INTO webhook_jobs (source_label, payload, status, created_at)
                VALUES (?, ?, 'pending', ?)
                """,
                (clean_text_value(source_label) or "/webhook", payload_json, created_at),
            )
            return int(cursor.lastrowid)


def claim_webhook_job(job_id: int) -> Optional[dict]:
    claimed_at = current_timestamp()
    with MONITOR_DB_LOCK:
        with get_monitor_db_connection() as conn:
            row = conn.execute(
                """
                SELECT id, source_label, payload, status, attempts
                FROM webhook_jobs
                WHERE id = ?
                LIMIT 1
                """,
                (job_id,),
            ).fetchone()
            if not row:
                return None

            job = dict(row)
            if job.get("status") == "done":
                return None

            conn.execute(
                """
                UPDATE webhook_jobs
                SET status = 'processing',
                    claimed_at = ?,
                    attempts = attempts + 1
                WHERE id = ?
                """,
                (claimed_at, job_id),
            )
            job["claimed_at"] = claimed_at
            job["attempts"] = int(job.get("attempts", 0) or 0) + 1
            return job


def complete_webhook_job(job_id: int) -> None:
    with MONITOR_DB_LOCK:
        with get_monitor_db_connection() as conn:
            conn.execute(
                """
                UPDATE webhook_jobs
                SET status = 'done',
                    completed_at = ?,
                    last_error = ''
                WHERE id = ?
                """,
                (current_timestamp(), job_id),
            )
    prune_webhook_jobs()


def fail_webhook_job(job_id: int, error_text: str) -> None:
    with MONITOR_DB_LOCK:
        with get_monitor_db_connection() as conn:
            conn.execute(
                """
                UPDATE webhook_jobs
                SET status = 'failed',
                    last_error = ?,
                    completed_at = ?
                WHERE id = ?
                """,
                (clean_text_value(error_text), current_timestamp(), job_id),
            )
    prune_webhook_jobs()


def prune_webhook_jobs() -> None:
    with MONITOR_DB_LOCK:
        with get_monitor_db_connection() as conn:
            conn.execute(
                """
                DELETE FROM webhook_jobs
                WHERE status IN ('done', 'failed')
                  AND id NOT IN (
                    SELECT id
                    FROM webhook_jobs
                    WHERE status IN ('done', 'failed')
                    ORDER BY id DESC
                    LIMIT ?
                )
                """,
                (WEBHOOK_JOB_HISTORY_LIMIT,),
            )


def get_recoverable_webhook_job_ids(limit: int = 1000) -> List[int]:
    with MONITOR_DB_LOCK:
        with get_monitor_db_connection() as conn:
            rows = conn.execute(
                """
                SELECT id
                FROM webhook_jobs
                WHERE status IN ('pending', 'processing')
                ORDER BY id ASC
                LIMIT ?
                """,
                (limit,),
            ).fetchall()
    return [int(row["id"] if isinstance(row, sqlite3.Row) else row[0]) for row in rows]


def requeue_pending_webhook_jobs() -> None:
    pending_job_ids = get_recoverable_webhook_job_ids()
    if not pending_job_ids:
        return

    ensure_webhook_worker()
    for job_id in pending_job_ids:
        WEBHOOK_QUEUE.put(job_id)

    logger.warning("WEBHOOK RECOVERY -> requeued_jobs=%s", len(pending_job_ids))
    add_monitor_event(
        direction="system",
        event_type="webhook_recovery",
        body=f"{len(pending_job_ids)} jobs requeued",
        status="ok",
        detail="Recovered pending webhook jobs from DB",
    )


def compute_whatsapp_signature(raw_body: bytes) -> str:
    digest = hmac.new(
        WHATSAPP_APP_SECRET.encode("utf-8"),
        raw_body,
        hashlib.sha256,
    ).hexdigest()
    return f"sha256={digest}"


def verify_whatsapp_signature(raw_body: bytes) -> Tuple[bool, str]:
    if not WHATSAPP_APP_SECRET:
        return True, "skipped"

    provided_signature = clean_text_value(request.headers.get("X-Hub-Signature-256"))
    if not provided_signature:
        return False, "missing_signature"

    expected_signature = compute_whatsapp_signature(raw_body)
    if not hmac.compare_digest(expected_signature, provided_signature):
        return False, "invalid_signature"

    return True, "verified"


def parse_incoming_webhook_request(source_label: str) -> Tuple[Optional[dict], Optional[Response], str]:
    raw_body = request.get_data(cache=True) or b""
    is_valid_signature, signature_reason = verify_whatsapp_signature(raw_body)
    if not is_valid_signature:
        add_monitor_event(
            direction="system",
            event_type="webhook_rejected",
            body=source_label,
            status="rejected",
            detail=signature_reason,
        )
        logger.warning("WEBHOOK REJECTED -> source=%s reason=%s", source_label, signature_reason)
        return None, jsonify({"ok": False, "error": signature_reason}), signature_reason

    payload = request.get_json(silent=True) or {}
    return payload, None, signature_reason


def is_monitor_login_enabled() -> bool:
    return bool(MONITOR_USERNAME and MONITOR_PASSWORD)


def has_monitor_session() -> bool:
    return bool(session.get("monitor_auth"))


def is_valid_monitor_credentials(username: str, password: str) -> bool:
    if not is_monitor_login_enabled():
        return False
    return hmac.compare_digest(clean_text_value(username), MONITOR_USERNAME) and hmac.compare_digest(
        clean_text_value(password),
        MONITOR_PASSWORD,
    )


def redirect_to_monitor_login():
    return redirect(url_for("monitor_login", next=request.full_path if request.query_string else request.path))


def is_safe_local_redirect_target(target: str) -> bool:
    cleaned = clean_text_value(target)
    if not cleaned:
        return False

    target_parts = urlparse(cleaned)
    if target_parts.scheme or target_parts.netloc:
        return False
    if not cleaned.startswith("/"):
        return False
    return True


def resolve_monitor_next_url(raw_target: Optional[str]) -> str:
    cleaned = clean_text_value(raw_target)
    if is_safe_local_redirect_target(cleaned):
        return cleaned
    return url_for("monitor_dashboard")


def is_monitor_authorized() -> bool:
    if has_monitor_session():
        return True

    if is_monitor_login_enabled():
        return False

    if not MONITOR_TOKEN:
        return True

    token = (
        request.args.get("token")
        or request.form.get("token")
        or request.headers.get("X-Monitor-Token")
    )
    return token == MONITOR_TOKEN


def get_monitor_cors_origin() -> str:
    origin = normalize_origin_value(request.headers.get("Origin") or "")
    if not origin:
        return ""
    if "*" in MONITOR_CORS_ORIGINS:
        return "*"

    origin_parts = urlparse(origin)

    for allowed_origin in MONITOR_CORS_ORIGINS:
        if origin == allowed_origin:
            return origin
        if not allowed_origin.endswith(":*"):
            continue

        allowed_parts = urlparse(allowed_origin[:-2])
        if (
            origin_parts.scheme == allowed_parts.scheme
            and origin_parts.hostname == allowed_parts.hostname
        ):
            return origin

    return ""


def bootstrap_monitor_storage() -> None:
    global MONITOR_VERSION
    try:
        init_monitor_db()
        load_runtime_settings()
        db_version = get_max_monitor_version_from_db()
        with MONITOR_VERSION_COND:
            MONITOR_VERSION = max(MONITOR_VERSION, db_version)
        logger.info("Monitor DB listo path=%s version=%s", MONITOR_DB_PATH, MONITOR_VERSION)
    except Exception as e:
        logger.exception("No se pudo inicializar monitor DB: %s", e)


def webhook_worker_loop() -> None:
    while True:
        job_id = WEBHOOK_QUEUE.get()
        try:
            job = claim_webhook_job(int(job_id))
            if not job:
                continue

            payload = json.loads(job.get("payload") or "{}")
            process_webhook_payload(payload)
            complete_webhook_job(int(job["id"]))
        except Exception as e:
            logger.exception("Error procesando job_id=%s en worker: %s", job_id, e)
            fail_webhook_job(int(job_id), str(e))
            add_monitor_event(
                direction="system",
                event_type="webhook_exception",
                body=f"job_id={job_id}",
                status="exception",
                detail=str(e),
            )
        finally:
            WEBHOOK_QUEUE.task_done()


def ensure_webhook_worker() -> None:
    global WEBHOOK_WORKER_STARTED

    with WEBHOOK_WORKER_LOCK:
        if WEBHOOK_WORKER_STARTED:
            return

        worker = Thread(target=webhook_worker_loop, name="whatsapp-webhook-worker", daemon=True)
        worker.start()
        WEBHOOK_WORKER_STARTED = True
        logger.info("WEBHOOK WORKER started")


def dispatch_webhook_payload(payload: dict, source_label: str) -> str:
    job_id: Optional[int] = None
    try:
        job_id = create_webhook_job(source_label, payload)
        ensure_webhook_worker()
        WEBHOOK_QUEUE.put(job_id)
        logger.info(
            "WEBHOOK ENQUEUED -> source=%s job_id=%s queue_size=%s",
            source_label,
            job_id,
            WEBHOOK_QUEUE.qsize(),
        )
        return "queued"
    except Exception as e:
        logger.exception("No se pudo encolar %s, procesando en linea: %s", source_label, e)
        add_monitor_event(
            direction="system",
            event_type="webhook_queue_fallback",
            body=source_label,
            status="fallback",
            detail=str(e),
        )
        process_webhook_payload(payload)
        if job_id is not None:
            complete_webhook_job(job_id)
        return "inline"


def extract_text_from_message(message: dict) -> Optional[str]:
    msg_type = message.get("type")

    if msg_type == "text":
        return optional_text_value(message.get("text", {}).get("body"))

    if msg_type == "interactive":
        interactive = message.get("interactive", {})
        button_reply = interactive.get("button_reply", {})
        list_reply = interactive.get("list_reply", {})

        # ✅ Primero ID (para flujos)
        return optional_text_value(
            button_reply.get("id")
            or list_reply.get("id")
            or button_reply.get("title")
            or list_reply.get("title")
        )

    return None

def send_whatsapp_interactive(to_phone: str, interactive: dict) -> None:
    to_phone = clean_text_value(to_phone)
    body_text = optional_text_value((interactive or {}).get("body", {}).get("text"))
    if not to_phone:
        add_monitor_event(
            direction="system",
            event_type="send_interactive_skipped",
            body="interactive",
            status="skipped",
            detail="missing recipient",
        )
        logger.warning("SKIP SEND INTERACTIVE -> missing recipient")
        return
    if not body_text:
        add_monitor_event(
            direction="system",
            event_type="send_interactive_skipped",
            peer=to_phone,
            body="interactive",
            status="skipped",
            detail="missing body.text",
        )
        logger.warning("SKIP SEND INTERACTIVE -> to=%s missing body.text", to_phone)
        return

    url = f"https://graph.facebook.com/{WHATSAPP_API_VERSION}/{WHATSAPP_PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {WHATSAPP_ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": to_phone,
        "type": "interactive",
        "interactive": interactive,
    }

    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=20)
        logger.info("SEND INTERACTIVE -> to=%s status=%s", to_phone, resp.status_code)
        add_monitor_event(
            direction="outbound",
            event_type="interactive",
            peer=to_phone,
            body=str(interactive),
            status="ok" if resp.status_code < 300 else "error",
            detail=f"http={resp.status_code}",
        )
        if resp.status_code >= 300:
            logger.error("GRAPH ERROR (interactive): %s", resp.text)
    except Exception as e:
        add_monitor_event(
            direction="outbound",
            event_type="interactive",
            peer=to_phone,
            body=str(interactive),
            status="exception",
            detail=str(e),
        )
        logger.exception("EXCEPTION sending interactive to=%s err=%s", to_phone, e)

# ========= Spring connector =========
def is_spring_auth_enabled() -> bool:
    return bool(SPRING_AUTH_BODY_JSON)


def clear_spring_auth_token() -> None:
    global SPRING_AUTH_TOKEN
    with SPRING_AUTH_LOCK:
        SPRING_AUTH_TOKEN = ""


def load_spring_auth_body() -> Optional[dict]:
    if not SPRING_AUTH_BODY_JSON:
        return None
    try:
        parsed = json.loads(SPRING_AUTH_BODY_JSON)
    except json.JSONDecodeError as exc:
        logger.error("SPRING AUTH BODY JSON invalido: %s", exc)
        return None
    if not isinstance(parsed, dict):
        logger.error("SPRING AUTH BODY JSON debe ser un objeto JSON")
        return None
    return parsed


def extract_json_path_value(payload: object, path: str) -> Optional[object]:
    current = payload
    for segment in [part for part in clean_text_value(path).split(".") if part]:
        if not isinstance(current, dict):
            return None
        current = current.get(segment)
    return current


def fetch_spring_auth_token(force_refresh: bool = False) -> Optional[str]:
    global SPRING_AUTH_TOKEN

    if not is_spring_auth_enabled():
        return None

    with SPRING_AUTH_LOCK:
        if SPRING_AUTH_TOKEN and not force_refresh:
            return SPRING_AUTH_TOKEN

        auth_body = load_spring_auth_body()
        if not auth_body:
            return None

        try:
            response = requests.post(
                SPRING_AUTH_URL,
                json=auth_body,
                timeout=10,
            )
        except Exception as exc:
            logger.exception("SPRING AUTH EXCEPTION: %s", exc)
            return None

        if response.status_code >= 300:
            logger.error("SPRING AUTH ERROR %s - %s", response.status_code, response.text)
            return None

        try:
            data = response.json() if response.content else {}
        except Exception as exc:
            logger.exception("SPRING AUTH JSON invalido: %s", exc)
            return None

        token_value = clean_text_value(extract_json_path_value(data, SPRING_AUTH_TOKEN_JSON_PATH))
        if not token_value:
            logger.error("SPRING AUTH sin token usable en path=%s", SPRING_AUTH_TOKEN_JSON_PATH)
            return None

        SPRING_AUTH_TOKEN = token_value
        return SPRING_AUTH_TOKEN


def build_spring_request_headers(force_refresh: bool = False) -> dict:
    headers = {"Content-Type": "application/json"}
    if not is_spring_auth_enabled():
        return headers

    token_value = fetch_spring_auth_token(force_refresh=force_refresh)
    if not token_value:
        return headers

    headers[SPRING_AUTH_HEADER_NAME] = f"{SPRING_AUTH_HEADER_PREFIX}{token_value}"
    return headers


def call_spring(
    sender: str,
    text: str,
    msg_id: str,
    timestamp: str,
    phone_number_id: str,
    is_new_session: bool = False,
) -> dict:
    """
    Flask -> Spring
    Espera respuesta:
      {"actions":[{"type":"text","body":"..."}, {"type":"template","name":"...","lang":"es_CO","params":[...]}]}
    """
    url = f"{SPRING_BASE_URL}/api/chatbot/inbound"
    payload = {
        "from": sender,
        "text": text,
        "messageId": msg_id,
        "timestamp": timestamp,
        "phoneNumberId": phone_number_id,
        "newConversation": is_new_session,
    }

    try:
        headers = build_spring_request_headers()
        r = requests.post(url, json=payload, headers=headers, timeout=10)
        if r.status_code in {401, 403} and is_spring_auth_enabled():
            clear_spring_auth_token()
            headers = build_spring_request_headers(force_refresh=True)
            r = requests.post(url, json=payload, headers=headers, timeout=10)

        if r.status_code >= 300:
            logger.error("SPRING ERROR %s - %s", r.status_code, r.text)
            add_monitor_event(
                direction="system",
                event_type="spring_error",
                peer=sender,
                body=text,
                status="error",
                detail=f"http={r.status_code}",
            )
            return {"actions": [{"type": "text", "body": API_UNAVAILABLE_MESSAGE}]}
        return r.json() if r.content else {"actions": []}
    except Exception as e:
        logger.exception("EXCEPTION calling Spring: %s", e)
        add_monitor_event(
            direction="system",
            event_type="spring_exception",
            peer=sender,
            body=text,
            status="exception",
            detail=str(e),
        )
        return {"actions": [{"type": "text", "body": API_UNAVAILABLE_MESSAGE}]}


# ========= WhatsApp senders =========
def send_whatsapp_text(to_phone: str, body: str) -> None:
    to_phone = clean_text_value(to_phone)
    body = clean_text_value(body)
    if not to_phone:
        add_monitor_event(
            direction="system",
            event_type="send_text_skipped",
            body="text",
            status="skipped",
            detail="missing recipient",
        )
        logger.warning("SKIP SEND TEXT -> missing recipient")
        return
    if not body:
        add_monitor_event(
            direction="system",
            event_type="send_text_skipped",
            peer=to_phone,
            body="text",
            status="skipped",
            detail="empty body",
        )
        logger.warning("SKIP SEND TEXT -> to=%s empty body", to_phone)
        return

    url = f"https://graph.facebook.com/{WHATSAPP_API_VERSION}/{WHATSAPP_PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {WHATSAPP_ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": to_phone,
        "type": "text",
        "text": {"body": body},
    }

    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=20)
        logger.info("SEND TEXT -> to=%s status=%s", to_phone, resp.status_code)
        add_monitor_event(
            direction="outbound",
            event_type="text",
            peer=to_phone,
            body=body,
            status="ok" if resp.status_code < 300 else "error",
            detail=f"http={resp.status_code}",
        )
        if resp.status_code >= 300:
            logger.error("GRAPH ERROR (text): %s", resp.text)
    except Exception as e:
        add_monitor_event(
            direction="outbound",
            event_type="text",
            peer=to_phone,
            body=body,
            status="exception",
            detail=str(e),
        )
        logger.exception("EXCEPTION sending text to=%s err=%s", to_phone, e)


def send_whatsapp_template(
    to_phone: str,
    template_name: str,
    language_code: str = "es_CO",
    body_params: Optional[List[str]] = None,
) -> None:
    to_phone = clean_text_value(to_phone)
    template_name = clean_text_value(template_name)
    filtered_body_params = [
        param_text
        for param in (body_params or [])
        if (param_text := clean_text_value(param))
    ]
    if not to_phone:
        add_monitor_event(
            direction="system",
            event_type="send_template_skipped",
            body="template",
            status="skipped",
            detail="missing recipient",
        )
        logger.warning("SKIP SEND TEMPLATE -> missing recipient")
        return
    if not template_name:
        add_monitor_event(
            direction="system",
            event_type="send_template_skipped",
            peer=to_phone,
            body="template",
            status="skipped",
            detail="missing template name",
        )
        logger.warning("SKIP SEND TEMPLATE -> to=%s missing template name", to_phone)
        return

    url = f"https://graph.facebook.com/{WHATSAPP_API_VERSION}/{WHATSAPP_PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {WHATSAPP_ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }

    components = []
    if filtered_body_params:
        components.append({
            "type": "body",
            "parameters": [{"type": "text", "text": p} for p in filtered_body_params],
        })

    payload = {
        "messaging_product": "whatsapp",
        "to": to_phone,
        "type": "template",
        "template": {
            "name": template_name,
            "language": {"code": language_code},
            **({"components": components} if components else {}),
        },
    }

    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=20)
        logger.info("SEND TEMPLATE -> to=%s status=%s", to_phone, resp.status_code)
        add_monitor_event(
            direction="outbound",
            event_type="template",
            peer=to_phone,
            body=template_name,
            status="ok" if resp.status_code < 300 else "error",
            detail=f"http={resp.status_code}",
        )
        if resp.status_code >= 300:
            logger.error("GRAPH ERROR (template): %s", resp.text)
    except Exception as e:
        add_monitor_event(
            direction="outbound",
            event_type="template",
            peer=to_phone,
            body=template_name,
            status="exception",
            detail=str(e),
        )
        logger.exception("EXCEPTION sending template to=%s err=%s", to_phone, e)


def send_whatsapp_image(to_phone: str, image_url: str, caption: Optional[str] = None) -> None:
    to_phone = clean_text_value(to_phone)
    image_url = clean_text_value(image_url)
    caption = optional_text_value(caption)
    if not to_phone:
        add_monitor_event(
            direction="system",
            event_type="send_image_skipped",
            body="image",
            status="skipped",
            detail="missing recipient",
        )
        logger.warning("SKIP SEND IMAGE -> missing recipient")
        return
    if not image_url:
        add_monitor_event(
            direction="system",
            event_type="send_image_skipped",
            peer=to_phone,
            body="image",
            status="skipped",
            detail="missing image url",
        )
        logger.warning("SKIP SEND IMAGE -> to=%s missing image url", to_phone)
        return

    url = f"https://graph.facebook.com/{WHATSAPP_API_VERSION}/{WHATSAPP_PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {WHATSAPP_ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }
    image_payload = {"link": image_url}
    if caption:
        image_payload["caption"] = caption

    payload = {
        "messaging_product": "whatsapp",
        "to": to_phone,
        "type": "image",
        "image": image_payload,
    }

    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=20)
        logger.info("SEND IMAGE -> to=%s status=%s", to_phone, resp.status_code)
        add_monitor_event(
            direction="outbound",
            event_type="image",
            peer=to_phone,
            body=image_url,
            status="ok" if resp.status_code < 300 else "error",
            detail=f"http={resp.status_code}",
        )
        if resp.status_code >= 300:
            logger.error("GRAPH ERROR (image): %s", resp.text)
    except Exception as e:
        add_monitor_event(
            direction="outbound",
            event_type="image",
            peer=to_phone,
            body=image_url,
            status="exception",
            detail=str(e),
        )
        logger.exception("EXCEPTION sending image to=%s err=%s", to_phone, e)


# ========= Webhook parsing =========
def iterate_incoming_messages(payload: dict):
    """
    Solo devuelve mensajes reales (value.messages).
    También loguea cambios sin messages (statuses, etc.) para debug.
    """
    for entry in payload.get("entry", []):
        for change in entry.get("changes", []):
            field = change.get("field")
            value = change.get("value", {}) or {}
            msgs = value.get("messages", []) or []

            logger.info("CHANGE field=%s has_messages=%s", field, bool(msgs))

            for msg in msgs:
                yield msg


def extract_phone_number_id(payload: dict) -> str:
    """
    Saca el phone_number_id real desde metadata del payload (si viene).
    """
    try:
        return (
            payload.get("entry", [{}])[0]
            .get("changes", [{}])[0]
            .get("value", {})
            .get("metadata", {})
            .get("phone_number_id", "")
        ) or ""
    except Exception:
        return ""


def dedupe_msg_id(msg_id: Optional[str]) -> bool:
    """
    True => ya lo procesamos (saltar)
    False => nuevo (procesar)
    """
    if not msg_id:
        return False

    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        with MONITOR_DB_LOCK:
            with get_monitor_db_connection() as conn:
                row = conn.execute(
                    "SELECT 1 FROM processed_messages WHERE msg_id = ? LIMIT 1",
                    (msg_id,),
                ).fetchone()
                if row:
                    return True

                conn.execute(
                    "INSERT INTO processed_messages (msg_id, created_at) VALUES (?, ?)",
                    (msg_id, created_at),
                )
                conn.execute(
                    """
                    DELETE FROM processed_messages
                    WHERE msg_id NOT IN (
                        SELECT msg_id
                        FROM processed_messages
                        ORDER BY created_at DESC, msg_id DESC
                        LIMIT ?
                    )
                    """,
                    (MAX_PROCESSED_CACHE,),
                )
                return False
    except Exception as e:
        logger.exception("Error usando deduplicacion persistente, usando memoria: %s", e)

    if msg_id in PROCESSED_MSG_IDS:
        return True

    PROCESSED_MSG_IDS.add(msg_id)
    if len(PROCESSED_MSG_IDS) > MAX_PROCESSED_CACHE:
        PROCESSED_MSG_IDS.clear()
    return False


def purge_expired_sessions(now: datetime) -> None:
    expired = []

    for sender, session in list(SESSIONS.items()):
        idle_time = now - session.last_activity
        if idle_time > SESSION_TIMEOUT:
            expired.append((sender, int(idle_time.total_seconds())))

    for sender, idle_seconds in expired:
        del SESSIONS[sender]
        logger.info("SESSION EXPIRED -> sender=%s idle_seconds=%s", sender, idle_seconds)
        add_monitor_event(
            direction="system",
            event_type="session_timeout",
            peer=sender,
            body="Session expired after inactivity",
            status="ok",
            detail=f"idle_seconds={idle_seconds}",
        )

        
def get_or_create_session(sender: str, now: datetime) -> Tuple[SessionState, bool]:
    """
    Retorna (session, is_new) donde is_new=True si es una nueva sesión real.
    """
    session = SESSIONS.get(sender)

    if session is None:
        session = SessionState(step="started")
        session.last_activity = now
        SESSIONS[sender] = session
        logger.info("SESSION CREATED -> sender=%s", sender)
        return session, True

    session.last_activity = now
    return session, False

def process_webhook_payload(payload: dict) -> None:
    global BOT_PAUSED
    phone_number_id = extract_phone_number_id(payload)
    now = datetime.now()
    purge_expired_sessions(now)

    for message in iterate_incoming_messages(payload):
        msg_id = message.get("id")
        sender = message.get("from")
        msg_type = message.get("type")
        incoming_text = extract_text_from_message(message)
        monitor_body = summarize_message_for_monitor(msg_type, incoming_text)
        ts = message.get("timestamp", "")

        if dedupe_msg_id(msg_id):
            logger.info("DEDUP msg_id=%s", msg_id)
            add_monitor_event(
                direction="inbound",
                event_type=msg_type or "unknown",
                peer=sender or "-",
                body=monitor_body,
                status="dedup",
                detail=f"msg_id={msg_id or '-'}",
            )
            continue

        logger.info("INCOMING: from=%s type=%s msg_id=%s text=%s", sender, msg_type, msg_id, incoming_text)
        add_monitor_event(
            direction="inbound",
            event_type=msg_type or "unknown",
            peer=sender or "-",
            body=monitor_body,
            status="ok",
            detail=f"msg_id={msg_id or '-'}",
        )

        if not sender:
            continue

        if not incoming_text:
            ignored_event_type = "empty_inbound_ignored" if msg_type == "text" else "unsupported_inbound_ignored"
            logger.info(
                "IGNORED INBOUND WITHOUT TEXT -> from=%s type=%s msg_id=%s",
                sender,
                msg_type,
                msg_id,
            )
            add_monitor_event(
                direction="system",
                event_type=ignored_event_type,
                peer=sender,
                body=f"[{clean_text_value(msg_type) or 'unknown'}]",
                status="ignored",
                detail=f"msg_id={msg_id or '-'}",
            )
            continue

        if BOT_PAUSED:
            add_monitor_event(
                direction="system",
                event_type="paused_block",
                peer=sender,
                body=incoming_text,
                status="blocked",
                detail="Mensaje procesado con BOT_PAUSED=true",
            )
            send_whatsapp_text(sender, BOT_PAUSED_MESSAGE)
            continue

        _, is_first_interaction = get_or_create_session(sender, now)

        if is_first_interaction and CHATBOT_ENROLLMENT_WELCOME_IMAGE_URL:
            send_whatsapp_image(
                to_phone=sender,
                image_url=CHATBOT_ENROLLMENT_WELCOME_IMAGE_URL,
                caption=CHATBOT_ENROLLMENT_WELCOME_IMAGE_CAPTION or None,
            )

        # ✅ Primera vez + template (opcional)
        if is_first_interaction and USE_WELCOME_TEMPLATE:
            send_whatsapp_template(
                to_phone=sender,
                template_name=WELCOME_TEMPLATE_NAME,
                language_code=WELCOME_TEMPLATE_LANG,
                body_params=["RutaDigital HARO"],
            )
            continue

        # ✅ Preguntar a Spring qué responder
        spring_out = call_spring(
            sender,
            incoming_text,
            msg_id or "",
            ts,
            phone_number_id,
            is_new_session=is_first_interaction,
        )
        actions = spring_out.get("actions", []) or []

        # Evita mostrar "opción no reconocida" en el primer mensaje.
        if is_first_interaction and actions:
            first = actions[0]
            first_type = (first.get("type") or "").lower()
            first_body = first.get("body", "")

            if first_type == "text" and is_unrecognized_menu_warning(first_body):
                if len(actions) > 1:
                    actions = actions[1:]
                    logger.info("FIRST_MESSAGE_FILTER sender=%s removed_warning=true", sender)
                else:
                    logger.info("FIRST_MESSAGE_FILTER sender=%s fallback=menu", sender)
                    menu_out = call_spring(sender, "menu", msg_id or "", ts, phone_number_id)
                    menu_actions = menu_out.get("actions", []) or []
                    if menu_actions:
                        actions = menu_actions

        logger.info("SPRING actions=%s", len(actions))

        for action in actions:
            a_type = (action.get("type") or "").lower()

            if a_type == "text":
                action_body = clean_text_value(action.get("body"))
                if not action_body:
                    record_skipped_action(sender, a_type, "empty body from Spring")
                    continue
                send_whatsapp_text(sender, action_body)

            elif a_type == "template":
                template_name = clean_text_value(action.get("name"))
                if not template_name:
                    record_skipped_action(sender, a_type, "missing template name from Spring")
                    continue
                send_whatsapp_template(
                    to_phone=sender,
                    template_name=template_name,
                    language_code=action.get("lang", "es_CO"),
                    body_params=action.get("params", []) or [],
                )

            elif a_type == "interactive_list":
                data = action.get("data") or {}
                action_body = clean_text_value(action.get("body"))
                sections = data.get("sections") or []
                if not action_body:
                    record_skipped_action(sender, a_type, "missing body from Spring")
                    continue
                if not isinstance(sections, list) or not sections:
                    record_skipped_action(sender, a_type, "missing sections from Spring")
                    continue
                interactive = {
                    "type": "list",
                    "body": {"text": action_body},
                    "action": {
                        "button": data.get("buttonText", "Ver opciones"),
                        "sections": sections,
                    },
                }
                send_whatsapp_interactive(sender, interactive)

            elif a_type == "interactive_buttons":
                data = action.get("data") or {}
                action_body = clean_text_value(action.get("body"))
                buttons = data.get("buttons") or []
                if not action_body:
                    record_skipped_action(sender, a_type, "missing body from Spring")
                    continue
                if not isinstance(buttons, list) or not buttons:
                    record_skipped_action(sender, a_type, "missing buttons from Spring")
                    continue
                interactive = {
                    "type": "button",
                    "body": {"text": action_body},
                    "action": {
                        "buttons": buttons,
                    },
                }
                send_whatsapp_interactive(sender, interactive)

            else:
                send_whatsapp_text(sender, "⚠️ Acción no soportada aún. Escribe *menu*.")

bootstrap_monitor_storage()
ensure_webhook_worker()
requeue_pending_webhook_jobs()

# ========= Routes =========
@app.get("/")
def home():
    if is_monitor_login_enabled() and not has_monitor_session():
        return redirect(url_for("monitor_login"))
    if MONITOR_TOKEN and not is_monitor_login_enabled():
        return "Monitor protegido. Abre /monitor?token=TU_MONITOR_TOKEN", 200
    return redirect(url_for("monitor_dashboard"))


@app.route("/login", methods=["GET", "POST"])
def monitor_login():
    if not is_monitor_login_enabled():
        return redirect(url_for("monitor_dashboard"))

    if has_monitor_session():
        return redirect(url_for("monitor_dashboard"))

    error_message = ""
    next_url = resolve_monitor_next_url(request.args.get("next") or request.form.get("next"))

    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        if is_valid_monitor_credentials(username, password):
            session["monitor_auth"] = True
            return redirect(next_url)
        error_message = "Credenciales invalidas."

    return render_template("login.html", error_message=error_message, next_url=next_url)


@app.post("/logout")
def monitor_logout():
    session.pop("monitor_auth", None)
    if is_monitor_login_enabled():
        return redirect(url_for("monitor_login"))
    return redirect(url_for("monitor_dashboard"))


@app.get("/health")
def health():
    return jsonify({"ok": True, "service": "whatsapp-webhook", "version": WHATSAPP_API_VERSION}), 200


@app.get("/webhook")
def verify_webhook():
    mode = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")

    if not mode and not token and not challenge:
        return (
            "Webhook activo. Este endpoint espera hub.mode, hub.verify_token y hub.challenge desde Meta.",
            200,
        )

    if mode == "subscribe" and token == WHATSAPP_VERIFY_TOKEN:
        return challenge or "", 200
    return "verification failed", 403


@app.get("/api/whatsapp/webhook")
def verify_webhook_api():
    return verify_webhook()


@app.after_request
def apply_monitor_cors_headers(response):
    if request.path in MONITOR_CORS_PATHS:
        allow_origin = get_monitor_cors_origin()
        if allow_origin:
            response.headers["Access-Control-Allow-Origin"] = allow_origin
            response.headers["Vary"] = "Origin"
            response.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
            response.headers["Access-Control-Allow-Headers"] = "Content-Type, X-Monitor-Token"
            response.headers["Access-Control-Expose-Headers"] = "Content-Disposition"
    return response


@app.route("/monitor/events", methods=["OPTIONS"])
@app.route("/api/monitor/events", methods=["OPTIONS"])
@app.route("/monitor/pause", methods=["OPTIONS"])
@app.route("/api/monitor/pause", methods=["OPTIONS"])
@app.route("/monitor/stream", methods=["OPTIONS"])
@app.route("/api/monitor/stream", methods=["OPTIONS"])
@app.route("/monitor/export/contacts.xlsx", methods=["OPTIONS"])
@app.route("/api/monitor/export/contacts.xlsx", methods=["OPTIONS"])
def monitor_cors_preflight():
    return "", 204


@app.get("/monitor/events")
@app.get("/api/monitor/events")
def monitor_events():
    if not is_monitor_authorized():
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    try:
        since = int((request.args.get("since") or "0").strip())
    except ValueError:
        since = 0

    peer = (request.args.get("peer") or "").strip()
    peer = peer if peer else None

    events = get_monitor_events(since=since, peer=peer)
    return jsonify(
        {
            "ok": True,
            "paused": BOT_PAUSED,
            "count": len(events),
            "total": get_monitor_total_events(peer=peer),
            "version": get_monitor_version(),
            "peer": peer or "",
            "events": events,
        }
    ), 200


@app.get("/monitor/stream")
@app.get("/api/monitor/stream")
def monitor_stream():
    if not is_monitor_authorized():
        return "unauthorized", 401

    try:
        since = int((request.args.get("since") or "0").strip())
    except ValueError:
        since = 0

    def event_stream():
        nonlocal since
        yield "retry: 3000\n\n"
        while True:
            with MONITOR_VERSION_COND:
                if MONITOR_VERSION <= since:
                    MONITOR_VERSION_COND.wait(timeout=20)
                current_version = MONITOR_VERSION

            if current_version > since:
                payload = json.dumps({"version": current_version}, ensure_ascii=False)
                yield f"id: {current_version}\nevent: monitor_update\ndata: {payload}\n\n"
                since = current_version
            else:
                # Mantiene viva la conexion en proxies intermedios.
                yield "event: keepalive\ndata: {}\n\n"

    response = Response(event_stream(), mimetype="text/event-stream")
    response.headers["Cache-Control"] = "no-cache"
    response.headers["X-Accel-Buffering"] = "no"
    return response


@app.get("/monitor")
def monitor_dashboard():
    if not is_monitor_authorized():
        if is_monitor_login_enabled():
            return redirect_to_monitor_login()
        return "unauthorized", 401

    token = request.args.get("token", "")
    return render_template(
        "monitor.html",
        token=token,
        poll_seconds=MONITOR_REFRESH_SECONDS,
        events_url=url_for("monitor_events"),
        pause_url=url_for("monitor_pause"),
        stream_url=url_for("monitor_stream"),
        export_url=url_for("monitor_export_contacts"),
        monitor_login_enabled=is_monitor_login_enabled(),
    )


@app.get("/monitor/export/contacts.xlsx")
@app.get("/api/monitor/export/contacts.xlsx")
def monitor_export_contacts():
    if not is_monitor_authorized():
        return "unauthorized", 401

    try:
        rows = get_monitor_contact_rows()
        output = build_contacts_excel_file(rows)
    except RuntimeError as e:
        logger.error("No se pudo exportar contactos a Excel: %s", e)
        return jsonify({"ok": False, "error": str(e)}), 500
    except Exception as e:
        logger.exception("Error generando Excel de contactos: %s", e)
        return jsonify({"ok": False, "error": "export_failed"}), 500

    filename = f"numeros_bot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/monitor/pause")
@app.post("/api/monitor/pause")
def monitor_pause():
    if not is_monitor_authorized():
        return "unauthorized", 401

    global BOT_PAUSED
    payload = request.get_json(silent=True) or {}
    action = (
        payload.get("action")
        or request.form.get("action")
        or request.args.get("action")
        or ""
    ).strip().lower()
    if action == "pause":
        BOT_PAUSED = True
    elif action == "resume":
        BOT_PAUSED = False
    else:
        BOT_PAUSED = not BOT_PAUSED

    save_monitor_setting("BOT_PAUSED", "true" if BOT_PAUSED else "false")
    logger.warning("BOT_PAUSED changed to %s", BOT_PAUSED)
    add_monitor_event(
        direction="system",
        event_type="pause_toggle",
        status="ok",
        detail=f"BOT_PAUSED={BOT_PAUSED}",
    )

    if request.is_json:
        return jsonify({"ok": True, "paused": BOT_PAUSED}), 200

    token = request.form.get("token") or request.args.get("token", "")
    if token:
        return redirect(url_for("monitor_dashboard", token=token))
    return redirect(url_for("monitor_dashboard"))


@app.post("/webhook")
def receive_webhook():
    payload, error_response, signature_mode = parse_incoming_webhook_request("/webhook")
    if error_response is not None:
        return error_response, 403
    logger.info("POST /webhook keys=%s", list(payload.keys()))
    try:
        mode = dispatch_webhook_payload(payload, "/webhook")
    except Exception as e:
        logger.exception("Error procesando /webhook: %s", e)
        add_monitor_event(
            direction="system",
            event_type="webhook_exception",
            body="/webhook",
            status="exception",
            detail=str(e),
        )
        return jsonify({"ok": False}), 500
    return jsonify({"ok": True, "mode": mode, "signature": signature_mode}), 200


@app.post("/api/whatsapp/webhook")
def receive_webhook_api():
    payload, error_response, signature_mode = parse_incoming_webhook_request("/api/whatsapp/webhook")
    if error_response is not None:
        return error_response, 403
    logger.info("POST /api/whatsapp/webhook keys=%s", list(payload.keys()))
    try:
        mode = dispatch_webhook_payload(payload, "/api/whatsapp/webhook")
    except Exception as e:
        logger.exception("Error procesando /api/whatsapp/webhook: %s", e)
        add_monitor_event(
            direction="system",
            event_type="webhook_exception",
            body="/api/whatsapp/webhook",
            status="exception",
            detail=str(e),
        )
        return jsonify({"ok": False}), 500
    return jsonify({"ok": True, "mode": mode, "signature": signature_mode}), 200


# ========= Main =========
if __name__ == "__main__":
    ok, missing_vars = validate_env()
    if not ok:
        print("Faltan variables de entorno para ejecutar el bot:")
        for var in missing_vars:
            print(f"- {var}")
        raise SystemExit(1)

    port = int(os.getenv("PORT", "8000"))
    app.run(host="0.0.0.0", port=port, debug=True)
